"""generate_report.py — Daily HTML container tracking report for Fantasia.

Queries fantasia-tracking (APP_DB_URL) + fantasia-data (DATA_DB_URL),
evaluates exceptions, and writes a styled HTML report to _reports/.

Run: py -3.12 sync/generate_report.py
"""
from __future__ import annotations

import datetime as dt
import difflib
import os
from collections import defaultdict
from pathlib import Path

import psycopg
from psycopg.rows import dict_row
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Paths & config
# ---------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent.parent
load_dotenv(HERE / ".env", override=True)

DATA_DB_URL = os.environ["DATA_DB_URL"]
APP_DB_URL  = os.environ["SUPABASE_DB_URL"]

TODAY       = dt.date.today()
REPORTS_DIR = HERE / "_reports"

# ---------------------------------------------------------------------------
# Destination mapping
# ---------------------------------------------------------------------------

DESTINATION_ORDER = ["FH", "FW", "TS", "NJ", "CU", "CV", "CH", "OB"]

DESTINATION_LABELS = {
    "FH":    "Freight Horse",
    "FW":    "Fulfillment World",
    "TS":    "Taylored Edison",
    "NJ":    "NY/NJ Port",
    "CU":    "PDC Prime",
    "CV":    "PDC MCP",
    "CH":    "Charleston",
    "OB":    "FOB Port",
    "OTHER": "Other / Unknown",
}

SHIPVIA_DEST = {
    "DFH": "FH", "BFH": "FH",
    "BFW": "FW",
    "BTS": "TS", "DTS": "TS",
    "BNJ": "NJ",
    "BCU": "CU",
    "BCV": "CV",
    "BCH": "CH",
    "BOB": "OB",
}

# ---------------------------------------------------------------------------
# Queries
# ---------------------------------------------------------------------------

CONTAINERS_SQL = """
SELECT container_id, status, vessel, voyage, current_vessel, carrier, bol_number,
       import_number, ship_via, pod, eta, etd, map_token, shipsgo_tracking_id,
       is_transshipment, ts_port, route_legs, current_leg, closed_at, last_synced_at
FROM containers
ORDER BY eta ASC NULLS LAST
"""

ETA_SLIPPED_SQL = """
SELECT DISTINCT container_id FROM container_events WHERE event_type = 'ETA_UPDATE'
"""

INTERNAL_ETA_SQL = """
SELECT
    purchase_container_id           AS container_id,
    time_date                       AS record_eta,
    purchase_ship_via               AS ship_via,
    purchase_container_size         AS container_size,
    purchase_supplier_inv_number    AS inv_number,
    purchase_supplier_bol_number    AS bol_number,
    purchase_vessel_name            AS erp_vessel,
    purchase_voyage_number          AS erp_voyage,
    purchase_po_number              AS po_number,
    purchase_supplier_reference     AS order_ref,
    purchase_agent                  AS agent
FROM visibility
WHERE purchase_container_id IS NOT NULL AND purchase_container_id <> ''
ORDER BY purchase_container_id, purchase_agent, purchase_supplier_reference, purchase_po_number
"""

# ---------------------------------------------------------------------------
# Data fetch
# ---------------------------------------------------------------------------

def fetch_data() -> tuple[list[dict], set[str], dict[str, dict]]:
    with psycopg.connect(APP_DB_URL, row_factory=dict_row) as conn:
        with conn.cursor() as cur:
            cur.execute(CONTAINERS_SQL)
            containers = cur.fetchall()
        with conn.cursor() as cur:
            cur.execute(ETA_SLIPPED_SQL)
            eta_slipped = {r["container_id"] for r in cur.fetchall()}

    with psycopg.connect(DATA_DB_URL, row_factory=dict_row) as conn:
        with conn.cursor() as cur:
            cur.execute(INTERNAL_ETA_SQL)
            vis_rows = cur.fetchall()

    ref_map:  dict = defaultdict(lambda: defaultdict(list))  # {cid: {(agent_word, ref): [po]}}
    meta_map: dict = {}

    for row in vis_rows:
        cid        = row["container_id"]
        po         = row["po_number"] or ""
        ref        = row["order_ref"] or ""
        agent_word = (row["agent"] or "").split()[0] if row["agent"] else ""
        inv        = row["inv_number"] or ""

        if po:
            ref_map[cid][(agent_word, ref)].append(po)

        if cid not in meta_map:
            meta_map[cid] = {
                "internal_eta":   row["record_eta"],
                "ship_via":       row["ship_via"],
                "container_size": row["container_size"],
                "inv_numbers":    set(),
                "bol_number":     row["bol_number"] or None,
                "erp_vessel":     row["erp_vessel"] or None,
                "erp_voyage":     row["erp_voyage"] or None,
            }
        if inv:
            meta_map[cid]["inv_numbers"].add(inv)
        if not meta_map[cid]["bol_number"] and row["bol_number"]:
            meta_map[cid]["bol_number"] = row["bol_number"]
        if not meta_map[cid]["erp_vessel"] and row["erp_vessel"]:
            meta_map[cid]["erp_vessel"] = row["erp_vessel"]
        if not meta_map[cid]["erp_voyage"] and row["erp_voyage"]:
            meta_map[cid]["erp_voyage"] = row["erp_voyage"]
        # keep the latest ETA across rows for this container
        if row["record_eta"] and (
            not meta_map[cid]["internal_eta"]
            or row["record_eta"] > meta_map[cid]["internal_eta"]
        ):
            meta_map[cid]["internal_eta"] = row["record_eta"]

    for cid in meta_map:
        meta_map[cid]["inv_numbers"] = ", ".join(sorted(meta_map[cid]["inv_numbers"]))

    def build_po_refs(cid: str) -> str:
        groups = ref_map.get(cid, {})
        parts: list[str] = []
        plain_pos: list[str] = []
        for (agent, ref) in sorted(groups):
            pos = ", ".join(sorted(set(groups[(agent, ref)])))
            if agent or ref:
                prefix = " | ".join(filter(None, [agent, ref]))
                parts.append(f"[{prefix} | {pos}]")
            else:
                plain_pos.append(pos)
        if plain_pos:
            parts.append(", ".join(plain_pos))
        return ", ".join(parts) if parts else "—"

    internal = {cid: {**meta_map[cid], "po_refs": build_po_refs(cid)} for cid in meta_map}

    return containers, eta_slipped, internal


# ---------------------------------------------------------------------------
# Date normalisation
# ---------------------------------------------------------------------------

def _to_date(v) -> dt.date | None:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return dt.date.fromisoformat(str(v)[:10])
    except Exception:
        return None


def _is_business_day_within(closed_at, n: int = 5) -> bool:
    """True if closed_at is within n business days (Mon–Fri) of today."""
    d = _to_date(closed_at)
    if d is None:
        return False
    bd = 0
    check = d + dt.timedelta(days=1)
    while check <= TODAY:
        if check.weekday() < 5:     # Mon=0 … Fri=4
            bd += 1
        check += dt.timedelta(days=1)
    return bd <= n


# ---------------------------------------------------------------------------
# Exception evaluation
# ---------------------------------------------------------------------------

_INACTIVE   = frozenset({"ARRIVED", "DISCHARGED"})
_NO_ETA_EXC = frozenset({"ARRIVED", "DISCHARGED", "UNTRACKED"})
_PRE_DEP    = frozenset({"BOOKED", "LOADED"})          # triggers STATUS_MISMATCH
_NOT_LOADED = frozenset({"INPROGRESS", "NEW"})         # triggers DEPARTED_NOT_LOADED


def _norm_vessel(s):
    if not s: return ""
    s = s.upper().strip()
    for prefix in ("M/V ", "MV/ ", "MV ", "S/S ", "SS "):
        if s.startswith(prefix):
            s = s[len(prefix):]
    return " ".join(s.split())


def _norm_voyage(s):
    if not s: return ""
    return s.upper().strip().rstrip("WENS")


def vessel_changed(c, erp_vessel, erp_voyage):
    sg_vessel = _norm_vessel(c.get("vessel") or "")
    sg_voyage = _norm_voyage(c.get("voyage") or "")
    ev  = _norm_vessel(erp_vessel or "")
    evo = _norm_voyage(erp_voyage or "")
    if not sg_vessel or not ev:
        return False
    vessel_match = difflib.SequenceMatcher(None, sg_vessel, ev).ratio() >= 0.82
    voyage_match = (not sg_voyage or not evo or sg_voyage == evo or
                    sg_voyage.startswith(evo) or evo.startswith(sg_voyage))
    return not vessel_match and not voyage_match


def evaluate_exceptions(c: dict, eta_slipped: set, internal: dict) -> list[dict]:
    """Return [{code, severity, detail}] for this container (all that apply)."""
    cid     = c["container_id"]
    eta     = _to_date(c.get("eta"))
    status  = (c.get("status") or "").upper()
    int_eta = _to_date((internal.get(cid) or {}).get("internal_eta"))
    excs: list[dict] = []

    # OVERDUE
    if eta and status not in _NO_ETA_EXC and TODAY > eta:
        d = (TODAY - eta).days
        excs.append({"code": "OVERDUE", "severity": "CRITICAL",
                     "detail": f"{d} day{'s' if d != 1 else ''} overdue"})

    # DELAYED_VS_PLAN — port ETA + 4d drayage would be late at warehouse
    # Trigger when (SG ETA − internal ETA) >= −3 (i.e. warehouse late by ≥ 1 day)
    if eta and int_eta and status not in _INACTIVE:
        d = (eta - int_eta).days
        if d >= -3:
            excs.append({"code": "DELAYED_VS_PLAN", "severity": "WARNING",
                         "detail": str(d)})   # raw delta; _cause_text adds drayage

    # ETA_SLIPPED — not applicable once arrived
    if cid in eta_slipped and status not in _INACTIVE:
        excs.append({"code": "ETA_SLIPPED", "severity": "WARNING",
                     "detail": "ETA shifted"})

    # DEPARTED_NOT_LOADED
    etd = _to_date(c.get("etd"))
    if status in _NOT_LOADED and etd and TODAY > etd:
        excs.append({"code": "DEPARTED_NOT_LOADED", "severity": "WARNING",
                     "detail": "Vessel departed — not loaded"})

    # VESSEL_CHANGED
    if status not in _INACTIVE:
        ir = internal.get(cid) or {}
        if vessel_changed(c, ir.get("erp_vessel"), ir.get("erp_voyage")):
            excs.append({"code": "VESSEL_CHANGED", "severity": "WARNING",
                         "detail": f"SG:{c.get('vessel')} vs ERP:{ir.get('erp_vessel')}"})

    return excs


def _group_exceptions(members: list) -> list[dict]:
    """Group-level exceptions injected into every container in the group."""
    # Exclude ARRIVED/DISCHARGED from STATUS_MISMATCH evaluation
    active_statuses = {(c.get("status") or "").upper() for c, _ in members
                       if (c.get("status") or "").upper() not in _INACTIVE}
    excs = []
    if "SAILING" in active_statuses and (active_statuses & _PRE_DEP):
        excs.append({"code": "STATUS_MISMATCH", "severity": "WARNING",
                     "detail": "Mixed SAILING/BOOKED on same vessel"})
    return excs


# ---------------------------------------------------------------------------
# Sort helpers
# ---------------------------------------------------------------------------

_MAX_DATE = dt.date.max


def _container_sort_key(item: tuple) -> tuple:
    """CRITICAL → WARNING → non-exception (by ETA asc, container_id)."""
    c, excs = item
    codes = {e["code"] for e in excs}
    sevs  = {e["severity"] for e in excs}
    eta   = _to_date(c.get("eta")) or _MAX_DATE
    cid   = c.get("container_id") or ""

    if "OVERDUE" in codes:  return (0, eta, cid)
    if "WARNING" in sevs:   return (1, eta, cid)
    return (2, eta, cid)


def _exc_group_order(item: tuple) -> tuple:
    """Exception groups: CRITICAL first → WARNING; secondary earliest ETA."""
    key, members = item
    sevs = {e["severity"] for _, excs in members for e in excs}
    etas = [e for e in (_to_date(c.get("eta")) for c, _ in members) if e is not None]
    earliest = min(etas) if etas else _MAX_DATE
    name = key[0] if key else ""
    if "CRITICAL" in sevs:
        return (0, earliest, name)
    return (1, earliest, name)


def _group_status_category(members: list) -> int:
    """Classify a clean vessel group: 0=SAILING, 1=BOOKED/pre-dep, 2=ARRIVED/DISCHARGED."""
    statuses = {(c.get("status") or "").upper() for c, _ in members}
    # RECEIVED counts as inactive for grouping purposes → ARRIVED/DISCHARGED sub-group
    active = statuses - _INACTIVE - {"RECEIVED"}
    if not active:
        return 2  # all INACTIVE or RECEIVED
    if "SAILING" in active and not (active & _PRE_DEP):
        return 0  # active containers are all SAILING (may also include UNTRACKED)
    return 1      # no SAILING, or BOOKED/pre-dep present


def _clean_group_order(item: tuple) -> tuple:
    """Clean groups: by status category, then earliest ETA; Pending Assignment last in its category."""
    key, members = item
    cat  = _group_status_category(members)
    etas = [e for e in (_to_date(c.get("eta")) for c, _ in members) if e is not None]
    earliest = min(etas) if etas else _MAX_DATE
    if key is None:
        return (cat, 1, earliest, "")   # Pending at bottom of its sub-group
    return (cat, 0, earliest, key[0])


# ---------------------------------------------------------------------------
# Row colour
# ---------------------------------------------------------------------------

def _row_color(c: dict, excs: list[dict]) -> str:
    status = (c.get("status") or "").upper()
    codes  = {e["code"] for e in excs}
    sevs   = {e["severity"] for e in excs}
    if status == "RECEIVED": return "#f3e5f5"   # light purple
    if "OVERDUE" in codes:   return "#fde7e7"
    if status in _INACTIVE:  return "#f0f7f0"
    if "WARNING" in sevs:    return "#fff3e0"
    # Yellow highlight for containers arriving within 7 days (display only, not an exception)
    eta = _to_date(c.get("eta"))
    if eta and status not in _INACTIVE and 0 <= (eta - TODAY).days <= 7:
        return "#fffde7"
    return "#ffffff"


# ---------------------------------------------------------------------------
# HTML helpers
# ---------------------------------------------------------------------------

_F = "font-family:Calibri,Arial,sans-serif;"

_BADGE = {
    "SAILING":    "#1565c0",
    "ARRIVED":    "#2a7a2a",
    "DISCHARGED": "#2a7a2a",
    "UNTRACKED":  "#b07a00",
    "RECEIVED":   "#7b1fa2",   # purple — container closed in ERP, shown for 5 business days
}

_SHIPSGO_MAP = "https://map.shipsgo.com/ocean/shipments/{sid}?token={token}"

# Sub-group definitions for On-Time section: (category int, bg colour, label)
_SUBGROUPS = [
    (2, "#2e7d32", "ARRIVED / DISCHARGED"),
    (0, "#1565c0", "SAILING"),
    (1, "#455a64", "BOOKED / PRE-DEPARTURE"),
]


def _badge(status: str | None) -> str:
    s     = (status or "").upper()
    color = _BADGE.get(s, "#666")
    label = s or "—"
    return (
        f'<span style="background:{color};color:#fff;padding:2px 6px;border-radius:3px;'
        f'font-size:10px;font-weight:bold;{_F}white-space:nowrap">{label}</span>'
    )


def _fmt_date(v) -> str:
    d = _to_date(v)
    if d is None:
        return '<span style="color:#aaa">—</span>'
    return d.strftime("%d %b %Y")


def _fmt_date_short(v) -> str:
    d = _to_date(v)
    return d.strftime("%Y-%m-%d") if d else "—"


def _delta_html(eta_raw, int_eta_raw) -> str:
    grey = 'style="color:#aaa"'
    e, i = _to_date(eta_raw), _to_date(int_eta_raw)
    if e is None or i is None:
        return f'<span {grey}>—</span>'
    d = (e - i).days
    if d > 0:
        return f'<span style="color:#c62828;font-weight:bold">+{d}d</span>'
    if d < 0:
        return f'<span style="color:#2e7d32;font-weight:bold">{d}d</span>'
    return f'<span {grey}>0d</span>'


def _esc(s) -> str:
    if not s:
        return "—"
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def _route_breadcrumb(legs: list, current_leg: int) -> str:
    """Render journey breadcrumb HTML for vessel group header banners."""
    if not legs:
        return ""
    arrow = '<span style="color:rgba(255,255,255,0.4);margin:0 4px;">&rarr;</span>'
    parts: list[str] = []
    for i, node in enumerate(legs):
        if node["type"] == "port":
            name    = _esc(node.get("name") or "")
            label   = node.get("label") or ""
            raw_dt  = node.get("date")
            date_s  = ""
            if raw_dt:
                try:
                    d = dt.date.fromisoformat(str(raw_dt)[:10])
                    date_s = d.strftime("%b") + " " + str(d.day)
                except Exception:
                    pass
            sub = (label + " " + date_s).strip()
            inner = f'<span style="font-size:11px;color:rgba(255,255,255,0.65);">{name}</span>'
            if sub:
                inner += (f'<span style="font-size:10px;color:rgba(255,255,255,0.4);">'
                          f'{_esc(sub)}</span>')
            parts.append(
                f'<span style="display:inline-flex;flex-direction:column;'
                f'align-items:center;gap:1px;">{inner}</span>'
            )
        else:
            name   = _esc(node.get("name") or "")
            voyage = _esc(node.get("voyage") or "")
            label  = f"{name}&nbsp;{voyage}" if voyage else name
            if i == current_leg:
                parts.append(
                    f'<span style="font-size:11px;color:#FFD54F;font-weight:bold;'
                    f'font-style:italic;">{label}</span>'
                )
            else:
                parts.append(
                    f'<span style="font-size:11px;color:rgba(255,255,255,0.65);">{label}</span>'
                )
    return arrow.join(parts)


def _fmt_vessel(c: dict) -> str:
    """Format vessel for display: plain, or 'current → mother' when transshipped."""
    mother  = (c.get("vessel") or "").strip()
    current = (c.get("current_vessel") or "").strip()
    is_ts   = c.get("is_transshipment") or False
    if not mother:
        return "—"
    if is_ts and current and current.upper() != mother.upper():
        return f"{_esc(current)} &rarr; {_esc(mother)}"
    return _esc(mother)


def _fmt_vessel_plain(c: dict) -> str:
    """Plain-text vessel for Excel: 'current → mother' or just 'mother'."""
    mother  = (c.get("vessel") or "").strip()
    current = (c.get("current_vessel") or "").strip()
    is_ts   = c.get("is_transshipment") or False
    if is_ts and current and current.upper() != mother.upper():
        return f"{current} → {mother}"
    return mother


def _track_link(shipsgo_id: str | None, map_token: str | None) -> str:
    if not shipsgo_id or not map_token:
        return '<span style="color:#aaa">—</span>'
    url = _SHIPSGO_MAP.format(sid=shipsgo_id, token=map_token)
    return (
        f'<a href="{url}" target="_blank" '
        f'style="color:#1565c0;text-decoration:none;font-size:14px" '
        f'title="Track on Shipsgo">&#128279;</a>'
    )


def _cause_text(excs: list[dict]) -> str:
    """Human-readable cause string for the Cause column; returns HTML."""
    if not excs:
        return '<span style="color:#aaa">—</span>'
    parts: list[str] = []
    for exc in excs:
        code   = exc["code"]
        detail = exc["detail"]
        if code == "OVERDUE":
            parts.append("ETA passed &ndash; not arrived")
        elif code == "DELAYED_VS_PLAN":
            n_late = int(detail) + 4        # delta + 4d drayage = warehouse lateness
            parts.append(f"Port ETA +4d drayage = late by {n_late}d at warehouse")
        elif code == "ETA_SLIPPED":
            parts.append("ETA shifted from original")
        elif code == "DEPARTED_NOT_LOADED":
            parts.append("Vessel departed &ndash; not loaded")
        elif code == "STATUS_MISMATCH":
            parts.append("Mixed SAILING/BOOKED on same vessel")
        elif code == "VESSEL_CHANGED":
            parts.append("Vessel/Voyage changed vs ERP &mdash; possible bump")
    return " &middot; ".join(parts) if parts else '<span style="color:#aaa">—</span>'


def build_hbl_inv(ir: dict) -> str:
    bol = (ir.get("bol_number") or "").strip()
    inv = (ir.get("inv_numbers") or "").strip()
    parts = [p for p in [bol, inv] if p]
    return " &middot; ".join(_esc(p) for p in parts) if parts else "—"


def _inbound_via(c: dict, internal: dict) -> str:
    """Render 'import_number · ship_via [· container_size]' cell."""
    ir   = internal.get(c["container_id"]) or {}
    imp  = (c.get("import_number") or "").strip()
    sv   = (c.get("ship_via") or ir.get("ship_via") or "").strip()
    sz   = (ir.get("container_size") or "").strip()
    parts = [p for p in [imp, sv, sz] if p]
    return " &middot; ".join(_esc(p) for p in parts) if parts else "—"


# ---------------------------------------------------------------------------
# Shared vessel-group header text
# ---------------------------------------------------------------------------

def _vessel_hdr(group_key, members: list) -> str:
    n = len(members)
    if group_key is None:
        return f"Pending Assignment&nbsp;&nbsp;|&nbsp;&nbsp;{n}&nbsp;container{'s' if n != 1 else ''}"
    vessel_name, voyage = group_key
    etds     = [e for e in (_to_date(c.get("etd")) for c, _ in members) if e is not None]
    etas     = [e for e in (_to_date(c.get("eta")) for c, _ in members) if e is not None]
    carriers = [c.get("carrier") for c, _ in members if c.get("carrier")]
    first_c  = members[0][0] if members else {}
    parts    = [_fmt_vessel(first_c)]
    if voyage:
        parts.append(f"Voyage:&nbsp;{_esc(voyage)}")
    if carriers:
        parts.append(f"Carrier:&nbsp;{_esc(carriers[0])}")
    if etds:
        parts.append(f"ETD:&nbsp;{_fmt_date_short(min(etds))}")
    if etas:
        parts.append(f"ETA:&nbsp;{_fmt_date_short(max(etas))}")
    parts.append(f"{n}&nbsp;container{'s' if n != 1 else ''}")
    header_line = "&nbsp;&nbsp;|&nbsp;&nbsp;".join(parts)

    legs        = first_c.get("route_legs") or []
    current_leg = first_c.get("current_leg") or 0
    breadcrumb  = _route_breadcrumb(legs, current_leg)
    if breadcrumb:
        return (
            f'{header_line}<br>'
            f'<span style="font-weight:normal;font-size:10px;line-height:1.8;">'
            f'{breadcrumb}</span>'
        )
    return header_line


# ---------------------------------------------------------------------------
# HTML builder — shared row renderer
# ---------------------------------------------------------------------------

def _exc_table_rows(group_key, members, internal, w) -> None:
    """Render one vessel group's header + container rows into the Exceptions table."""
    members_sorted = sorted(members, key=_container_sort_key)
    w(f"""      <tr style="background:#7b1f1f;color:#fff;">
        <td colspan="11" style="padding:7px 8px;font-weight:bold;font-size:11px;{_F}">
          {_vessel_hdr(group_key, members_sorted)}
        </td>
      </tr>""")
    for c, excs in members_sorted:
        cid = c["container_id"]
        ir  = internal.get(cid) or {}
        w(f"""      <tr style="background:{_row_color(c, excs)};">
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;font-family:monospace">{_esc(cid)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444">{build_hbl_inv(ir)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444">{_inbound_via(c, internal)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8">{_badge(c.get('status'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444">{_esc(c.get('pod'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8">{_fmt_date(ir.get('internal_eta'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8">{_fmt_date(c.get('eta'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;text-align:center">{_delta_html(c.get('eta'), ir.get('internal_eta'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#555">{_esc(ir.get('po_refs'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444;font-style:italic">{_cause_text(excs)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;text-align:center">{_track_link(c.get('shipsgo_tracking_id'), c.get('map_token'))}</td>
      </tr>""")


def _ontime_table_rows(group_key, members, internal, w) -> None:
    """Render one vessel group's header + container rows into the On-Time table."""
    members_sorted = sorted(members, key=_container_sort_key)
    w(f"""      <tr style="background:#1a3c5e;color:#fff;">
        <td colspan="10" style="padding:7px 8px;font-weight:bold;font-size:11px;{_F}">
          {_vessel_hdr(group_key, members_sorted)}
        </td>
      </tr>""")
    for c, excs in members_sorted:
        cid = c["container_id"]
        ir  = internal.get(cid) or {}
        w(f"""      <tr style="background:{_row_color(c, excs)};">
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;font-family:monospace">{_esc(cid)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444">{build_hbl_inv(ir)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444">{_inbound_via(c, internal)}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8">{_badge(c.get('status'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#444">{_esc(c.get('pod'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8">{_fmt_date(ir.get('internal_eta'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8">{_fmt_date(c.get('eta'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;text-align:center">{_delta_html(c.get('eta'), ir.get('internal_eta'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;color:#555">{_esc(ir.get('po_refs'))}</td>
        <td style="padding:4px 6px;border-bottom:1px solid #e8e8e8;text-align:center">{_track_link(c.get('shipsgo_tracking_id'), c.get('map_token'))}</td>
      </tr>""")


# ---------------------------------------------------------------------------
# HTML builder
# ---------------------------------------------------------------------------

def build_html(containers: list[dict], eta_slipped: set, internal: dict) -> str:
    # Filter containers: drop those closed > 5 business days ago; override status for recent ones
    received_ids: set[str] = set()
    active_containers: list[dict] = []
    for c in containers:
        ca = c.get("closed_at")
        if ca is not None:
            if _is_business_day_within(ca):
                received_ids.add(c["container_id"])
                active_containers.append({**c, "status": "RECEIVED"})
            # else: closed > 5 business days ago — exclude from report
        else:
            active_containers.append(c)

    # Per-container exceptions; RECEIVED containers get empty list (excluded from all checks)
    enriched = [
        (c, [] if c["container_id"] in received_ids
            else evaluate_exceptions(c, eta_slipped, internal))
        for c in active_containers
    ]

    # Last-synced timestamp
    ls_vals = [c["last_synced_at"] for c in active_containers if c.get("last_synced_at")]
    last_synced = max(ls_vals) if ls_vals else None
    if last_synced and hasattr(last_synced, "strftime"):
        ls_str = last_synced.strftime("%d %b %Y %H:%M UTC")
    elif last_synced:
        ls_str = str(last_synced)[:16]
    else:
        ls_str = "—"

    # Assign each container to a destination bucket
    def _dest(c: dict) -> str:
        # Prefer ship_via on the containers row (available for closed containers too)
        sv = c.get("ship_via") or (internal.get(c["container_id"]) or {}).get("ship_via") or ""
        return SHIPVIA_DEST.get(sv.strip().upper(), "OTHER")

    dest_buckets: dict[str, list] = defaultdict(list)
    for item in enriched:
        dest_buckets[_dest(item[0])].append(item)

    out: list[str] = []
    w = out.append

    # ---- document shell ----
    w(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Fantasia - Container Tracking Daily Report</title>
</head>
<body style="margin:0;padding:16px;background:#f4f6f8;{_F}">
<div style="max-width:900px;margin:0 auto;background:#fff;border-radius:6px;
            overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.12);">""")

    # ---- main header ----
    w(f"""  <div style="background:#1a3c5e;padding:18px 24px;">
    <div style="color:#fff;font-size:18px;font-weight:bold;{_F}">
      Fantasia &mdash; Container Tracking Daily Report &mdash; {TODAY.strftime("%d %B %Y")}
    </div>
  </div>""")

    # ---- destination jump bar ----
    active_dests = [
        (dc, DESTINATION_LABELS.get(dc, "Other / Unknown"))
        for dc in DESTINATION_ORDER + ["OTHER"]
        if dest_buckets.get(dc)
    ]
    if active_dests:
        links = "&nbsp;&middot;&nbsp;".join(
            f'<a href="#dest-{dc}" style="color:#1a3c5e;text-decoration:none">{_esc(lbl)}</a>'
            for dc, lbl in active_dests
        )
        w(f"""  <div style="text-align:center;padding:8px 24px;font-size:12px;"""
          f"""color:#555;border-bottom:1px solid #e0e0e0;{_F}">"""
          f"""Jump to:&nbsp;{links}</div>""")

    # ===================================================================
    # Destination loop
    # ===================================================================
    for dest_code in DESTINATION_ORDER + ["OTHER"]:
        dest_items = dest_buckets.get(dest_code, [])
        if not dest_items:
            continue

        label = DESTINATION_LABELS.get(dest_code, "Other / Unknown")

        # Codes present in this destination's data (prefer containers.ship_via)
        codes_present = sorted({
            sv for c, _ in dest_items
            if (sv := (c.get("ship_via") or
                        (internal.get(c["container_id"]) or {}).get("ship_via") or
                        "").strip().upper())
        })
        codes_str = " &middot; ".join(_esc(cd) for cd in codes_present)

        # ---- destination banner ----
        banner_codes = f"&nbsp;&nbsp;({codes_str})" if codes_str else ""
        other_links = "&nbsp;&middot;&nbsp;".join(
            f'<a href="#dest-{dc}" style="color:#fff;text-decoration:underline;font-size:12px">'
            f'{_esc(dc)}</a>'
            for dc, _ in active_dests if dc != dest_code
        )
        w(f"""  <div id="dest-{dest_code}" style="background:#212121;padding:10px 24px;"""
          f"""margin-top:6px;display:table;width:100%;box-sizing:border-box;">
    <div style="display:table-cell;color:#fff;font-size:15px;font-weight:bold;{_F}">
      &#10022; {_esc(label.upper())} &mdash; {_esc(label)}{banner_codes}
    </div>
    <div style="display:table-cell;text-align:right;vertical-align:middle;white-space:nowrap">
      {other_links}
    </div>
  </div>""")

        # Build vessel groups scoped to this destination
        groups: dict = defaultdict(list)
        for c, excs in dest_items:
            vessel = (c.get("vessel") or "").strip() or None
            key    = (vessel, c.get("voyage")) if vessel else None
            groups[key].append((c, excs))

        # Inject group-level exceptions (STATUS_MISMATCH) per destination scope
        for key in list(groups.keys()):
            g_excs = _group_exceptions(groups[key])
            if g_excs:
                groups[key] = [(c, excs + g_excs) for c, excs in groups[key]]

        # Partition
        exc_groups:   list = []
        clean_groups: list = []
        for key, members in groups.items():
            if any(excs for _, excs in members):
                exc_groups.append((key, members))
            else:
                clean_groups.append((key, members))

        exc_groups   = sorted(exc_groups,   key=_exc_group_order)
        clean_groups = sorted(clean_groups, key=_clean_group_order)

        # Counts for this destination
        all_dest = [(c, excs) for _, members in groups.items() for c, excs in members]
        n_critical = sum(1 for _, e in all_dest if any(x["severity"] == "CRITICAL" for x in e))
        n_warning  = sum(1 for _, e in all_dest if any(x["severity"] == "WARNING"  for x in e))

        n_exc_groups       = len(exc_groups)
        n_exc_containers   = sum(len(m) for _, m in exc_groups)
        n_clean_groups     = len(clean_groups)
        n_clean_containers = sum(len(m) for _, m in clean_groups)

        cat_buckets: dict[int, list] = defaultdict(list)
        for kg in clean_groups:
            cat_buckets[_group_status_category(kg[1])].append(kg)

        # =================================================================
        # Exceptions section
        # =================================================================
        if exc_groups:
            w(f"""  <div style="background:#b71c1c;padding:8px 24px;">
    <div style="color:#fff;font-size:13px;font-weight:bold;{_F}">
      &#9888; EXCEPTIONS &mdash; {n_exc_groups}&nbsp;vessel group{'s' if n_exc_groups != 1 else ''},
      {n_exc_containers}&nbsp;container{'s' if n_exc_containers != 1 else ''}
    </div>
  </div>
  <div style="background:#fdf3f3;padding:6px 24px;font-size:12px;color:#555;{_F}">
    <span style="color:#b71c1c;font-weight:bold">{n_critical}&nbsp;critical</span>&nbsp;&middot;&nbsp;
    <span style="color:#e65100;font-weight:bold">{n_warning}&nbsp;warnings</span>
  </div>""")

            w(f"""  <div style="padding:0 24px 12px;">
    <table style="width:100%;border-collapse:collapse;font-size:10px;{_F}">
      <tr style="background:#1a3c5e;color:#fff;">
        <th style="padding:5px 6px;text-align:left;">Container</th>
        <th style="padding:5px 6px;text-align:left;">HBL&nbsp;/&nbsp;Inv#</th>
        <th style="padding:5px 6px;text-align:left;">Inbound&nbsp;#&nbsp;&middot;&nbsp;Via</th>
        <th style="padding:5px 6px;text-align:left;">Status</th>
        <th style="padding:5px 6px;text-align:left;">POD</th>
        <th style="padding:5px 6px;text-align:left;">Int.&nbsp;ETA</th>
        <th style="padding:5px 6px;text-align:left;">ETA</th>
        <th style="padding:5px 6px;text-align:center;">Delta</th>
        <th style="padding:5px 6px;text-align:left;">POs&nbsp;/&nbsp;Ref</th>
        <th style="padding:5px 6px;text-align:left;">Cause</th>
        <th style="padding:5px 6px;text-align:center;">Track</th>
      </tr>""")
            for group_key, members in exc_groups:
                _exc_table_rows(group_key, members, internal, w)
            w("    </table>\n  </div>")

        # =================================================================
        # On-Time section
        # =================================================================
        w(f"""  <div style="background:#1a3c5e;padding:8px 24px;">
    <div style="color:#fff;font-size:13px;font-weight:bold;{_F}">
      ON-TIME CONTAINERS &mdash; {n_clean_groups}&nbsp;vessel group{'s' if n_clean_groups != 1 else ''},
      {n_clean_containers}&nbsp;container{'s' if n_clean_containers != 1 else ''}
    </div>
  </div>
  <div style="background:#edf2f7;padding:6px 24px;font-size:12px;color:#555;{_F}">
    Sorted by earliest ETA&nbsp;&nbsp;&middot;&nbsp;&nbsp;Last synced:&nbsp;<strong>{ls_str}</strong>
  </div>""")

        w(f"""  <div style="padding:0 24px 12px;">
    <table style="width:100%;border-collapse:collapse;font-size:10px;{_F}">
      <tr style="background:#1a3c5e;color:#fff;">
        <th style="padding:5px 6px;text-align:left;">Container</th>
        <th style="padding:5px 6px;text-align:left;">HBL&nbsp;/&nbsp;Inv#</th>
        <th style="padding:5px 6px;text-align:left;">Inbound&nbsp;#&nbsp;&middot;&nbsp;Via</th>
        <th style="padding:5px 6px;text-align:left;">Status</th>
        <th style="padding:5px 6px;text-align:left;">POD</th>
        <th style="padding:5px 6px;text-align:left;">Int.&nbsp;ETA</th>
        <th style="padding:5px 6px;text-align:left;">ETA</th>
        <th style="padding:5px 6px;text-align:center;">Delta</th>
        <th style="padding:5px 6px;text-align:left;">POs&nbsp;/&nbsp;Ref</th>
        <th style="padding:5px 6px;text-align:center;">Track</th>
      </tr>""")
        for cat, sg_color, sg_label in _SUBGROUPS:
            sg = cat_buckets.get(cat, [])
            if not sg:
                continue
            n_sg = len(sg)
            w(f"""      <tr style="background:{sg_color};color:#fff;">
        <td colspan="10" style="padding:6px 8px;font-size:13px;font-weight:normal;{_F}">
          {sg_label} &mdash; {n_sg}&nbsp;vessel{'s' if n_sg != 1 else ''}
        </td>
      </tr>""")
            for group_key, members in sg:
                _ontime_table_rows(group_key, members, internal, w)
        w("    </table>\n  </div>")

    # ---- footer ----
    w(f"""  <div style="background:#f4f6f8;padding:10px 24px;text-align:center;font-size:11px;color:#888;{_F}">
    Generated automatically &middot; Source: fantasia-tracking + fantasia-data
  </div>""")

    w("</div>\n</body>\n</html>")
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_xlsx(containers: list[dict], eta_slipped: set, internal: dict):
    """Flat Excel workbook — one row per container, same order as HTML report."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Replicate filter + enrich from build_html
    received_ids: set[str] = set()
    active_containers: list[dict] = []
    for c in containers:
        ca = c.get("closed_at")
        if ca is not None:
            if _is_business_day_within(ca):
                received_ids.add(c["container_id"])
                active_containers.append({**c, "status": "RECEIVED"})
        else:
            active_containers.append(c)

    enriched = [
        (c, [] if c["container_id"] in received_ids
            else evaluate_exceptions(c, eta_slipped, internal))
        for c in active_containers
    ]

    def _dest(c: dict) -> str:
        sv = c.get("ship_via") or (internal.get(c["container_id"]) or {}).get("ship_via") or ""
        return SHIPVIA_DEST.get(sv.strip().upper(), "OTHER")

    dest_buckets: dict[str, list] = defaultdict(list)
    for item in enriched:
        dest_buckets[_dest(item[0])].append(item)

    # Build ordered flat row list matching HTML sort
    xl_rows: list[tuple] = []
    for dest_code in DESTINATION_ORDER + ["OTHER"]:
        dest_items = dest_buckets.get(dest_code, [])
        if not dest_items:
            continue
        dest_label = DESTINATION_LABELS.get(dest_code, "Other / Unknown")

        groups: dict = defaultdict(list)
        for c, excs in dest_items:
            vessel = (c.get("vessel") or "").strip() or None
            key    = (vessel, c.get("voyage")) if vessel else None
            groups[key].append((c, excs))
        for key in list(groups.keys()):
            g_excs = _group_exceptions(groups[key])
            if g_excs:
                groups[key] = [(c, excs + g_excs) for c, excs in groups[key]]

        exc_groups   = sorted([(k, m) for k, m in groups.items() if any(e for _, e in m)],  key=_exc_group_order)
        clean_groups = sorted([(k, m) for k, m in groups.items() if not any(e for _, e in m)], key=_clean_group_order)

        for _, members in exc_groups + clean_groups:
            for c, excs in sorted(members, key=_container_sort_key):
                xl_rows.append((dest_code, dest_label, c, excs, internal.get(c["container_id"]) or {}))

    # --- cell helpers ---
    def _strip_tz(val):
        if hasattr(val, "tzinfo") and val.tzinfo is not None:
            return val.replace(tzinfo=None)
        return val

    def _date_val(v):
        return _to_date(v)

    def _delta_val(eta_raw, int_eta_raw):
        e, i = _to_date(eta_raw), _to_date(int_eta_raw)
        return (e - i).days if e and i else None

    def _cause_plain(excs: list[dict]) -> str:
        parts: list[str] = []
        for exc in excs:
            code, detail = exc["code"], exc["detail"]
            if code == "OVERDUE":
                parts.append("ETA passed – not arrived")
            elif code == "DELAYED_VS_PLAN":
                parts.append(f"Port ETA +4d drayage = late by {int(detail) + 4}d at warehouse")
            elif code == "ETA_SLIPPED":
                parts.append("ETA shifted from original")
            elif code == "DEPARTED_NOT_LOADED":
                parts.append("Vessel departed – not loaded")
            elif code == "STATUS_MISMATCH":
                parts.append("Mixed SAILING/BOOKED on same vessel")
            elif code == "VESSEL_CHANGED":
                parts.append("Vessel/Voyage changed vs ERP — possible bump")
        return " · ".join(parts)

    def _row_fill(c: dict, excs: list[dict]) -> str | None:
        status = (c.get("status") or "").upper()
        codes  = {e["code"] for e in excs}
        sevs   = {e["severity"] for e in excs}
        if status == "RECEIVED":      return "F3E5F5"
        if "OVERDUE" in codes:        return "FDDEDE"
        if status in _INACTIVE:       return "F0F7F0"
        if "WARNING" in sevs:         return "FFF3E0"
        eta = _to_date(c.get("eta"))
        if eta and status not in _INACTIVE and 0 <= (eta - TODAY).days <= 7:
            return "FFFDE7"
        return None

    # --- workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Containers"

    HEADERS = [
        "Destination", "Destination Name", "Container", "Status",
        "Vessel", "Voyage", "Carrier", "POD",
        "Internal ETA", "Shipsgo ETA", "Delta (days)",
        "Inbound #", "Via", "Container Size",
        "HBL / Inv#", "POs / Ref", "Exceptions", "Cause",
        "Last Synced",
    ]
    COL_WIDTHS = [6, 18, 14, 12, 22, 10, 14, 8, 13, 13, 8, 10, 8, 10, 16, 42, 22, 40, 20]

    HDR_FILL = PatternFill("solid", fgColor="1A3C5E")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, col_idx)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.freeze_panes = "A2"

    for dest_code, dest_label, c, excs, ir in xl_rows:
        sv = (c.get("ship_via") or ir.get("ship_via") or "")
        ws.append([
            dest_code,
            dest_label,
            c.get("container_id") or "",
            c.get("status") or "",
            _fmt_vessel_plain(c),
            c.get("voyage") or "",
            c.get("carrier") or "",
            c.get("pod") or "",
            _strip_tz(_date_val(ir.get("internal_eta"))),
            _strip_tz(_date_val(c.get("eta"))),
            _delta_val(c.get("eta"), ir.get("internal_eta")),
            c.get("import_number") or "",
            sv,
            ir.get("container_size") or "",
            " · ".join(p for p in [ir.get("bol_number") or "", ir.get("inv_numbers") or ""] if p),
            ir.get("po_refs") or "",
            ", ".join(e["code"] for e in excs),
            _cause_plain(excs),
            _strip_tz(c.get("last_synced_at")),
        ])
        fill_hex = _row_fill(c, excs)
        if fill_hex:
            fill = PatternFill("solid", fgColor=fill_hex)
            row_idx = ws.max_row
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row_idx, col_idx).fill = fill

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Fetching data …")
    containers, eta_slipped, internal = fetch_data()
    print(f"  {len(containers)} containers · {len(eta_slipped)} ETA slips · {len(internal)} internal records")

    REPORTS_DIR.mkdir(exist_ok=True)

    html = build_html(containers, eta_slipped, internal)
    out_html = REPORTS_DIR / f"ContainerTracking_Report_{TODAY}.html"
    out_html.write_text(html, encoding="utf-8")
    print(f"HTML written   → {out_html}")

    excel_dir = HERE / "_excel"
    excel_dir.mkdir(exist_ok=True)
    wb = build_xlsx(containers, eta_slipped, internal)
    out_xlsx = excel_dir / f"ContainerTracking_Report_{TODAY:%Y-%m-%d}.xlsx"
    wb.save(out_xlsx)
    print(f"Excel written  → {out_xlsx}")


if __name__ == "__main__":
    main()
